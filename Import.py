import os
import sys
from openpyxl import load_workbook
from datetime import date

ROMANIAN_MONTHS = {
    "ianuarie": 1, "ian": 1, "ian.": 1,
    "februarie": 2, "feb": 2, "feb.": 2,
    "martie": 3, "mar": 3, "mar.": 3,
    "aprilie": 4, "apr": 4, "apr.": 4,
    "mai": 5,
    "iunie": 6, "iun": 6, "iun.": 6,
    "iulie": 7, "iul": 7, "iul.": 7,
    "august": 8, "aug": 8, "aug.": 8,
    "septembrie": 9, "sep": 9, "sept": 9, "sep.": 9, "sept.": 9,
    "octombrie": 10, "oct": 10, "oct.": 10,
    "noiembrie": 11, "noi": 11, "noi.": 11,
    "decembrie": 12, "dec": 12, "dec.": 12
}

def parse_month_to_number(luna):
    # acceptă număr sau denumire (română)
    if luna is None:
        return None
    s = str(luna).strip().lower()
    try:
        m = int(s)
        if 1 <= m <= 12:
            return m
    except ValueError:
        pass
    return ROMANIAN_MONTHS.get(s)

def determine_year_for_month(month_number):
    today = date.today()
    if month_number < today.month:
        return today.year + 1
    else:
        return today.year

def find_sheet_by_name(wb, name):
    if not name:
        return None
    target = name.strip().lower()
    for sheet in wb.sheetnames:
        if sheet.strip().lower() == target:
            return sheet
    return None

def main():
    # utilizare: acceptă fie 3 argumente (fisier, luna, folder_output) -> anul este calculat automat
    # fie 4 argumente (fisier, luna, an, folder_output) -> anul este luat din argument
    if len(sys.argv) not in (4, 5):
        print("Utilizare:")
        print("  python Import.py <fisier_excel> <luna> <folder_output>           # anul calculat automat")
        print("  python Import.py <fisier_excel> <luna> <an> <folder_output>     # an furnizat manual")
        sys.exit(1)

    fisier_excel = sys.argv[1]  # poate fi cale absoluta
    luna_arg = sys.argv[2]
    if len(sys.argv) == 5:
        an = sys.argv[3]
        folder_output = sys.argv[4]
    else:
        folder_output = sys.argv[3]
        # calculăm anul pe baza lunii
        month_num = parse_month_to_number(luna_arg)
        if month_num is None:
            print(f"Eroare: luna '{luna_arg}' nu e recunoscută. Folosiți număr (1-12) sau denumire în română.")
            print(f"Valori recunoscute: {sorted(set(ROMANIAN_MONTHS.keys()))}")
            sys.exit(1)
        an = str(determine_year_for_month(month_num))

    # verificare existenta folder output
    if not os.path.exists(folder_output):
        os.makedirs(folder_output)

    # constuire cale completa pentru fisierul de output
    nume_fisier_output = f"onomastici_{luna_arg}_{an}.txt"
    cale_fisier_output = os.path.join(folder_output, nume_fisier_output)

    # deschidere fisier excel
    try:
        wb = load_workbook(filename=fisier_excel)
    except Exception as e:
        print(f"Eroare la deschiderea fisierului Excel: {e}")
        sys.exit(1)

    # verificare daca pagina exista (case-insensitive)
    sheet_name = find_sheet_by_name(wb, luna_arg)
    if not sheet_name:
        print(f"Eroare: nu exista pagina '{luna_arg}' în fișier.")
        print(f"Pagini disponibile: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]

    # citire date si scriere in fisier text
    try:
        with open(cale_fisier_output, 'w', encoding='utf-8') as f_out:
            for row in ws.iter_rows(min_row=2, values_only=True):  # ignoram header
                nume_prenume = row[0]
                data_nastere = row[1]
                functie = row[2] if len(row) > 2 and row[2] else ""  # daca lipseste, folosim string gol

                if nume_prenume and data_nastere:
                    # spargem "Nume Prenume" -> prenume si nume
                    parti = str(nume_prenume).split(' ', 1)
                    prenume = parti[1] if len(parti) > 1 else ""
                    nume = parti[0]
                    linie_output = f"{prenume} {nume} - {data_nastere} - {functie}\n"
                    f_out.write(linie_output)
    except Exception as e:
        print(f"Eroare la scrierea în fișierul de output: {e}")
        sys.exit(1)

    print(f"Datele au fost scrise cu succes în '{cale_fisier_output}'")


if __name__ == "__main__":
    main()
